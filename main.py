import openpyxl
from openpyxl import load_workbook
from igraph import *
import sys
import os

from PyQt5 import QtWidgets, QtCore
from PyQt5.QtWidgets import QApplication, QTableWidgetItem, QMessageBox, QWidget
from OmGTU import Ui_MainWindow
from Dialog import Ui_Dialog

file = ''
graph_file = ''
all_articles = []
crt_articles = []
ranked_sources = []
ranks = []
ves_ranks = dict
wb = ws = row_count = column_count = None
deg = deg_in = deg_out = auth = betw = huby = clos = clos_in = clos_out = False
art_dict_start = 0
art_dict_end = 500
src_dict_start = 0
src_dict_end = 20


def get_graph_path():
    global graph_file
    graph_file = QtWidgets.QFileDialog.getOpenFileName()[0]


def load_all_papers():
    global all_articles, ranks
    all_articles = [[] for _ in range(row_count - 1)]
    counter = 0
    for j in range(2, row_count + 1):
        for i in range(1, column_count + 1):
            all_articles[counter].append(ws.cell(row=j, column=i).value)
            ranks.append(0)
        counter += 1
    return all_articles, ranks


def range_sort(dict):
    range_sorted_dict = {}
    out = {}
    sorted_tuples = sorted(dict.items(), key=lambda item: (item[1]), reverse=True)
    sorted_dict = {k + 1: v for k, v in sorted_tuples}
    i = 1
    for k, v in sorted_dict.items():
        out.setdefault(v, []).append(k)
    for k, v in out.items():
        for vv in v:
            range_sorted_dict[vv] = i
        i += 1
    return range_sorted_dict


def profile_range_sort(sorted_dict):
    range_sorted_dict = {}
    out = {}
    i = 1
    for k, v in sorted_dict.items():
        out.setdefault(v, []).append(k)
    for k, v in out.items():
        for vv in v:
            range_sorted_dict[vv] = i
        i += 1
    return range_sorted_dict


def profile_dict_sort(dict):  # Сортировка списка статей по центральности и создание списка отсортированных строк
    sorted_dict = {}
    sorted_keys = sorted(dict, key=dict.get)  # [1, 3, 2]
    for w in sorted_keys:
        sorted_dict[w] = dict[w]
    return sorted_dict


def dict_sum(dict0, dict1):
    sum_dict = dict0.copy()
    for k, v in dict1.items():
        sum_dict[k] = sum_dict.get(k, 0) + v
    return sum_dict


def referativ():
    global crt_articles, ranks
    graph = Graph.Read_Pajek(graph_file)
    degree = range_sort(dict(enumerate(Graph.degree(graph, mode='out'))))
    closeness = range_sort(dict(enumerate(Graph.closeness(graph, mode='out'))))
    hub = range_sort(dict(enumerate(Graph.hub_score(graph))))
    profiles = profile_dict_sort(dict_sum(dict_sum(degree, closeness), hub))
    sorted_profiles = profile_range_sort(profiles)  # Ключ - номер статьи, значение - ранг
    rows = list(sorted_profiles.keys())
    ranks = list(sorted_profiles.values())
    crt_articles = [[] for _ in range(len(rows))]
    for i in range(len(rows)):
        crt_articles[i] = all_articles[rows[i] - 1]
    return crt_articles, ranks


def priznan():
    global crt_articles, ranks
    graph = Graph.Read_Pajek(graph_file)
    degree = range_sort(dict(enumerate(Graph.degree(graph, mode='in'))))
    closeness = range_sort(dict(enumerate(Graph.closeness(graph, mode='in'))))
    authority = range_sort(dict(enumerate(Graph.authority_score(graph))))
    profiles = profile_dict_sort(dict_sum(dict_sum(degree, closeness), authority))
    sorted_profiles = profile_range_sort(profiles)
    rows = list(sorted_profiles.keys())
    ranks = list(sorted_profiles.values())
    crt_articles = [[] for _ in range(len(rows))]
    for i in range(len(rows)):
        crt_articles[i] = all_articles[rows[i] - 1]
    return crt_articles, ranks


def vesomost():
    global crt_articles, ranks, ves_ranks
    graph = Graph.Read_Pajek(graph_file)
    degree = range_sort(dict(enumerate(Graph.degree(graph))))
    closeness = range_sort(dict(enumerate(Graph.closeness(graph))))
    betweenness = range_sort(dict(enumerate(Graph.betweenness(graph))))
    authority = range_sort(dict(enumerate(Graph.authority_score(graph))))
    hub = range_sort(dict(enumerate(Graph.hub_score(graph))))
    profiles = profile_dict_sort(dict_sum(dict_sum(dict_sum(dict_sum(degree, closeness), betweenness), authority), hub))
    ves_ranks = sorted_profiles = profile_range_sort(profiles)
    rows = list(sorted_profiles.keys())
    ranks = list(sorted_profiles.values())
    crt_articles = [[] for _ in range(len(rows))]
    for i in range(len(rows)):
        crt_articles[i] = all_articles[rows[i] - 1]
    return crt_articles, ranks


def constructor():
    graph = Graph.Read_Pajek(graph_file)
    profile = {}
    if huby:
        hub_centr = range_sort(dict(enumerate(Graph.hub_score(graph))))
        profile = dict_sum(profile, hub_centr)
    if auth:
        authority_centr = range_sort(dict(enumerate(Graph.authority_score(graph))))
        profile = dict_sum(profile, authority_centr)
    if betw:
        betweenness_centr = range_sort(dict(enumerate(Graph.betweenness(graph))))
        profile = dict_sum(profile, betweenness_centr)
    if deg:
        degree_centr = range_sort(dict(enumerate(Graph.degree(graph))))
        profile = dict_sum(profile, degree_centr)
    if deg_in:
        degree_in_centr = range_sort(dict(enumerate(Graph.degree(graph, mode='in'))))
        profile = dict_sum(profile, degree_in_centr)
    if deg_out:
        degree_out_centr = range_sort(dict(enumerate(Graph.degree(graph, mode='out'))))
        profile = dict_sum(profile, degree_out_centr)
    if clos:
        closeness_centr = range_sort(dict(enumerate(Graph.closeness(graph))))
        profile = dict_sum(profile, closeness_centr)
    if clos_in:
        closeness_in_centr = range_sort(dict(enumerate(Graph.closeness(graph, mode='in'))))
        profile = dict_sum(profile, closeness_in_centr)
    if clos_out:
        closeness_out_centr = range_sort(dict(enumerate(Graph.closeness(graph, mode='out'))))
        profile = dict_sum(profile, closeness_out_centr)
    profiles = profile_dict_sort(profile)
    sorted_profiles = profile_range_sort(profiles)
    rows = list(sorted_profiles.keys())
    ranki = list(sorted_profiles.values())
    crt_src = [[] for _ in range(len(rows))]
    for i in range(len(rows)):
        crt_src[i] = all_articles[rows[i] - 1]
    return crt_src, ranki


class MyWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.dialog = Dialog()

        # Добавление названий профилей в выдвигающийся список
        profiles = ['Без профиля', 'Исходящий', 'Входящий', 'Входящий/Исходящий']
        self.ui.comboBox.addItems(profiles)
        self.ui.comboBox.addItem('Тест')

        search_filter = ['Ключевое слово', 'Название', 'Автор', 'УДК']
        self.ui.comboBox_2.addItems(search_filter)

        self.ui.lineEdit.setPlaceholderText("Поиск по ключевым словам...")

        self.ui.pushButton.clicked.connect(self.search)
        self.ui.comboBox.currentTextChanged.connect(self.renew)  # Фильтрация таблицы по профилю
        self.ui.action_excel.triggered.connect(self.excel_save)
        self.ui.articles_load.triggered.connect(self.get_filepath)
        self.ui.network_load.triggered.connect(get_graph_path)
        self.ui.next_button.clicked.connect(self.get_next)
        self.ui.previous_button.clicked.connect(self.get_previous)
        self.ui.all_articles.clicked.connect(self.load_all_articles)
        self.ui.pushButton_2.clicked.connect(self.show_dialog)
        self.ui.constr_pushButton.clicked.connect(self.check_centr)

    def check_centr(self):
        global deg, deg_in, deg_out, auth, betw, huby, clos, clos_in, clos_out
        if self.ui.degree_checkBox.isChecked():
            deg = True
        else:
            deg = False
        if self.ui.deg_in_checkBox_7.isChecked():
            deg_in = True
        else:
            deg_in = False
        if self.ui.deg_out_checkBox_6.isChecked():
            deg_out = True
        else:
            deg_out = False
        if self.ui.auth_checkBox_3.isChecked():
            auth = True
        else:
            auth = False
        if self.ui.betw_checkBox_2.isChecked():
            betw = True
        else:
            betw = False
        if self.ui.hub_checkBox_4.isChecked():
            huby = True
        else:
            huby = False
        if self.ui.clos_checkBox_5.isChecked():
            clos = True
        else:
            clos = False
        if self.ui.clos_in_checkBox_9.isChecked():
            clos_in = True
        else:
            clos_in = False
        if self.ui.clos_out_checkBox_8.isChecked():
            clos_out = True
        else:
            clos_out = False
        mylist, ranks_list = constructor()
        self.printer(mylist, ranks_list)

    def show_dialog(self):
        self.get_sources()
        self.dialog.show()

    def load_all_articles(self):
        self.printer(crt_articles, ranks)

    def get_filepath(self):
        global file, crt_articles, ranks
        global wb, ws, row_count, column_count
        file = QtWidgets.QFileDialog.getOpenFileName()[0]
        wb = load_workbook(filename=file,
                           data_only=True)
        ws = wb.active
        row_count = ws.max_row
        column_count = ws.max_column
        crt_articles, ranks = load_all_papers()
        self.printer(crt_articles, ranks)

    def get_previous(self):
        global art_dict_start, art_dict_end
        self.ui.tableWidget.clear()
        art_dict_end = art_dict_start
        art_dict_start -= 500
        if art_dict_start < 0:
            art_dict_start = 0
        self.printer(crt_articles, ranks)

    def get_next(self):
        global art_dict_start, art_dict_end
        self.ui.tableWidget.clear()
        art_dict_start = art_dict_end
        art_dict_end += 500
        if art_dict_end > len(crt_articles):
            art_dict_end = len(crt_articles)
        self.printer(crt_articles, ranks)

    def printer(self, mylist: list, ranks):  # Вывод таблицы на экран, задача кол-ва строк и столбцов, их имён
        column_names = ['№ строки', 'Название', 'Авторы', 'УДК', 'Ключевые слова',
                        'Издание', 'Том, выпуск, № издания', 'Год', 'Страницы', 'Ссылка']
        self.ui.tableWidget.setColumnCount(column_count)  # Задача кол-ва столбцов и строк
        self.ui.tableWidget.setHorizontalHeaderLabels(column_names)
        self.ui.tableWidget.setRowCount(art_dict_end - art_dict_start)
        row = 0
        col = 0
        col_position = 1
        # Заполнение таблицы в приложении
        for i in range(art_dict_start, art_dict_end):
            for item in mylist[i]:
                self.ui.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(item)))
                col += 1
        self.ui.tableWidget.insertColumn(col_position)
        column_names.insert(1, 'Ранг')
        self.ui.tableWidget.setHorizontalHeaderLabels(column_names)
        for j in range(art_dict_start, art_dict_end):
            self.ui.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
            self.ui.tableWidget.setItem(row, col_position, QTableWidgetItem(str(ranks[j])))
            row += 1

    def search(self):
        search_result = []
        row = 0
        col = 0
        rows_count = 0
        text = self.ui.lineEdit.text()  # Считывание текста введённого пользователем
        items = self.ui.tableWidget.findItems(text, QtCore.Qt.MatchContains)  # Поиск совпадений
        #  Добавление их в список
        for item in items:
            if item and item.column() == self.search_renew():
                i = item.row()
                for j in range(0, 11):
                    search_result.append(self.ui.tableWidget.item(i, j).text())
                rows_count += 1  # запомнить количество найденных строк, для их вывода
        if search_result:
            #  Очистка таблицы и вывод только искомых данных
            self.ui.tableWidget.clear()
            self.ui.tableWidget.setColumnCount(11)  # Задача кол-ва столбцов и строк
            columns_headers = ['№ строки', 'Ранг', 'Название', 'Авторы', 'УДК', 'Ключевые слова',
                               'Издание', 'Том, выпуск, № издания', 'Год', 'Страницы', 'Ссылка']
            self.ui.tableWidget.setHorizontalHeaderLabels(columns_headers)
            self.ui.tableWidget.setRowCount(rows_count)
            for item in search_result:
                self.ui.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
                self.ui.tableWidget.setItem(row, col, QTableWidgetItem(str(item)))
                col += 1
        else:
            self.ui.tableWidget.clear()
            self.ui.tableWidget.setColumnCount(3)  # Задача кол-ва столбцов и строк
            self.ui.tableWidget.setHorizontalHeaderLabels(
                ('Ничего', 'Не', 'Найдено')
            )
            self.ui.tableWidget.setRowCount(0)

    def search_renew(self):  # Выбор сортировки
        checker = str(self.ui.comboBox_2.currentText())
        if checker == 'Ключевое слово':
            x = 5
        elif checker == 'Автор':
            x = 3
        elif checker == 'УДК':
            x = 4
        elif checker == 'Название':
            x = 2
        return x

    def renew(self):  # Выбор сортировки
        checker = str(self.ui.comboBox.currentText())
        if checker == 'Без профиля':
            global crt_articles
            crt_articles = all_articles
            self.printer(all_articles, ranks)
        if checker == 'Входящий/Исходящий':
            mylist, rank_list = vesomost()
            self.printer(mylist, rank_list)
        if checker == 'Исходящий':
            mylist, rank_list = referativ()
            self.printer(mylist, rank_list)
        if checker == 'Входящий':
            mylist, rank_list = priznan()
            self.printer(mylist, rank_list)

    def get_rows(self):
        columns_headers = ['№ строки', 'Ранг', 'Название', 'Авторы', 'УДК', 'Ключевые слова',
                           'Издание', 'Том, выпуск, № издания', 'Год', 'Страницы', 'Ссылка']
        rows_cnt = self.ui.tableWidget.rowCount()
        colums_cnt = self.ui.tableWidget.columnCount()
        rows = [[] for _ in range(rows_cnt)]
        for i in range(rows_cnt):
            for j in range(colums_cnt):
                rows[i].append(self.ui.tableWidget.item(i, j).text())
        rows.insert(0, columns_headers)
        return rows

    def excel_save(self):
        counter = 0
        swb = openpyxl.Workbook()
        sws = swb.active
        ext = 'xlsx'
        filename = '\Ваша выборка статей'
        basename = os.environ['USERPROFILE'] + '\Desktop' + filename
        actualname = "%s.%s" % (basename, ext)
        sws.title = 'Выбранные статьи'
        rows = self.get_rows()
        check_file = os.path.exists(actualname)
        for row in rows:
            sws.append(row)
        if check_file:
            counter += 1
            actualname = "%s (%d).%s" % (basename, counter, ext)
            swb.save(actualname)
            filename = filename.replace('\\', '') + ' (' + str(counter) + ').' + ext
        else:
            swb.save(actualname)
            filename = filename.replace('\\', '')
        QMessageBox.about(self, 'Где мой файл?', 'Ваш файл на рабочем столе\n'
                          + 'Имя файла: ' + filename)

    def get_sources(self):
        sources = {}
        global ranked_sources
        ranked_sources = [[] for _ in range(row_count + 1)]
        vesomost()
        cnt2 = 0
        for i in range(2, row_count + 1):
            if str(ws.cell(row=i, column=6).value) not in sources:
                sources[str(ws.cell(row=i, column=6).value)] = [1, ves_ranks.get(int(ws.cell(row=i, column=1).value))]
            elif str(ws.cell(row=i, column=6).value) in sources:
                (sources[str(ws.cell(row=i, column=6).value)])[0] += 1
                (sources[str(ws.cell(row=i, column=6).value)])[1] += ves_ranks.get(int(ws.cell(row=i, column=1).value))
        for key in sources.keys():
            ranked_sources[cnt2].append(str(key))
            for i in range(0, 2):
                ranked_sources[cnt2].append(str((sources[key])[i]))
            cnt2 += 1
        ranked_sources = [item for item in ranked_sources if item]
        self.dialog.printer(ranked_sources)


class Dialog(QWidget):
    def __init__(self):
        super().__init__()
        self.di = Ui_Dialog()
        self.di.setupUi(self)
        self.di.pushButton_5.clicked.connect(self.search)
        self.di.pushButton_6.clicked.connect(self.show_all_sources)
        self.di.pushButton.clicked.connect(self.get_next)
        self.di.pushButton_2.clicked.connect(self.get_previous)
        self.di.pushButton_4.clicked.connect(self.src_excel_save)

    def show_all_sources(self):
        self.printer(ranked_sources)

    def printer(self, mylist: list):  # Вывод таблицы на экран, задача кол-ва строк и столбцов, их имён
        column_names = ['Наименование издания', 'Кол-во статей', 'Ранг издания']
        self.di.tableWidget.setColumnCount(3)  # Задача кол-ва столбцов и строк
        self.di.tableWidget.setHorizontalHeaderLabels(column_names)
        self.di.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        self.di.tableWidget.setColumnWidth(0, 500)
        self.di.tableWidget.setRowCount(src_dict_end - src_dict_start)
        row = 0
        col = 0
        # Заполнение таблицы в приложении
        for i in range(src_dict_start, src_dict_end):
            for item in mylist[i]:
                self.di.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
                self.di.tableWidget.setItem(row, col, QTableWidgetItem(str(item)))
                col += 1

    def search(self):
        search_result = []
        row = 0
        col = 0
        rows_count = 0
        text = self.di.lineEdit.text()  # Считывание текста введённого пользователем
        items = self.di.tableWidget.findItems(text, QtCore.Qt.MatchContains)  # Поиск совпадений
        #  Добавление их в список
        for item in items:
            if item and item.column() == 0:
                i = item.row()
                for j in range(0, 3):
                    search_result.append(self.di.tableWidget.item(i, j).text())
                rows_count += 1  # запомнить количество найденных строк, для их вывода
        if search_result:
            #  Очистка таблицы и вывод только искомых данных
            self.di.tableWidget.clear()
            self.di.tableWidget.setColumnCount(3)  # Задача кол-ва столбцов и строк
            columns_headers = ['Наименование издания', 'Кол-во статей', 'Ранг издания']
            self.di.tableWidget.setHorizontalHeaderLabels(columns_headers)
            self.di.tableWidget.setRowCount(rows_count)
            for item in search_result:
                self.di.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
                self.di.tableWidget.setItem(row, col, QTableWidgetItem(str(item)))
                col += 1
        else:
            self.di.tableWidget.clear()
            self.di.tableWidget.setColumnCount(3)  # Задача кол-ва столбцов и строк
            self.di.tableWidget.setHorizontalHeaderLabels(
                ('Ничего', 'Не', 'Найдено')
            )
            self.di.tableWidget.setRowCount(0)

    def get_previous(self):
        global src_dict_start, src_dict_end
        self.di.tableWidget.clear()
        src_dict_end = src_dict_start
        src_dict_start -= 20
        if src_dict_start < 0:
            src_dict_start = 0
        self.printer(ranked_sources)

    def get_next(self):
        global src_dict_start, src_dict_end
        self.di.tableWidget.clear()
        src_dict_start = src_dict_end
        src_dict_end += 20
        if src_dict_end > len(crt_articles):
            src_dict_end = len(crt_articles)
        self.printer(ranked_sources)

    def get_src_rows(self):
        columns_headers = ['Наименование издания', 'Кол-во статей', 'Ранг издания']
        rows_cnt = self.di.tableWidget.rowCount()
        colums_cnt = self.di.tableWidget.columnCount()
        rows = [[] for _ in range(rows_cnt)]
        for i in range(rows_cnt):
            for j in range(colums_cnt):
                rows[i].append(self.di.tableWidget.item(i, j).text())
        rows.insert(0, columns_headers)
        return rows

    def src_excel_save(self):
        counter = 0
        srcswb = openpyxl.Workbook()
        srcsws = srcswb.active
        ext = 'xlsx'
        filename = '\Ваша выборка источников'
        basename = os.environ['USERPROFILE'] + '\Desktop' + filename
        actualname = "%s.%s" % (basename, ext)
        srcsws.title = 'Выбранные источники'
        rows = self.get_src_rows()
        check_file = os.path.exists(actualname)
        for row in rows:
            srcsws.append(row)
        if check_file:
            counter += 1
            actualname = "%s (%d).%s" % (basename, counter, ext)
            srcswb.save(actualname)
            filename = filename.replace('\\', '') + ' (' + str(counter) + ').' + ext
        else:
            srcswb.save(actualname)
            filename = filename.replace('\\', '')
        QMessageBox.about(self, 'Где мой файл?', 'Ваш файл на рабочем столе\n'
                          + 'Имя файла: ' + filename)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    application = MyWindow()
    application.show()
    sys.exit(app.exec())
